import os
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl.styles import Alignment, Border, Side

import re

def parse_event_data(event_folder):
    events = []
    for subdir in sorted(os.listdir(event_folder)):
        subdir_path = os.path.join(event_folder, subdir)
        if os.path.isdir(subdir_path):
            event_file = os.path.join(subdir_path, 'event.xml')
            if os.path.exists(event_file):
                tree = ET.parse(event_file)
                root = tree.getroot()
                
                event_id = root.find('name/id').text
                event_str = root.find('name/str').text
                
                # 移除名稱中的冗餘 ID 前綴 (例如 260122_01_1：)
                # 使用正則匹配 "數字_數字_數字：" 或 "數字_數字_數字 "
                clean_name = re.sub(r'^\d+_\d+_\d+[：: ]+', '', event_str).strip()
                
                if len(event_id) >= 9:
                    date_str = f'20{event_id[0:2]}-{event_id[2:4]}-{event_id[4:6]}'
                    month_str = f'{int(event_id[2:4])}月'
                    day_str = f'{int(event_id[2:4])}月{int(event_id[4:6])}日'
                    
                    event_type_code = event_id[6:8]
                    index = event_id[8]
                    
                    event_type_dict = {
                        "01": "01(介紹)",
                        "02": "02(歌曲解鎖)",
                        "03": "03(區域解鎖)",
                        "04": "04(區域解鎖)",
                        "05": "05(完美挑戰)",
                        "06": "06(票券配布)",
                        "07": "07(印章卡獎勵)",
                        "09": "09(傳導放寬)",
                        "11": "11(任務追加)"
                    }
                    event_type_name = event_type_dict.get(event_type_code, "未知類型")
                    
                    # 決定標籤與權重（用於同一天內容的排序）
                    content_label = ""
                    priority = 99
                    if event_type_code == "03" or "新區域" in clean_name or "ちほー 解禁" in clean_name:
                        content_label = "【新區域】"
                        priority = 1
                    elif event_type_code == "02" or "新曲" in clean_name or "追加" in clean_name:
                        content_label = "【新曲】"
                        priority = 2
                    elif event_type_code == "05":
                        content_label = "【活動】"
                        priority = 3
                    elif event_type_code == "01" and "介紹" in clean_name:
                        # 介紹通常是重複的資訊，標註為資訊或過濾
                        content_label = "【資訊】"
                        priority = 4
                    elif event_type_code == "09" or "系統" in clean_name:
                        content_label = "【系統】"
                        priority = 5
                    else:
                        content_label = "【其他】"
                        priority = 6
                    
                    events.append({
                        'id': event_id,
                        'name': clean_name,
                        'date': date_str,
                        'month': month_str,
                        'day': day_str,
                        'type_code': event_type_code,
                        'type_name': event_type_name,
                        'index': index,
                        'content_label': content_label,
                        'priority': priority
                    })
    return events

def export_to_txt(events, output_file, output_type):
    with open(output_file, 'w', encoding='utf-8') as f:
        for e in events:
            if output_type == '1':
                f.write(f"{e['id']} {e['name']}\n")
            elif output_type == '2':
                f.write(f"ID: {e['id']}\n")
                f.write(f"Name: {e['name']}\n")
                f.write(f"Date: {e['date']} ({e['day']})\n")
                f.write(f"Type: {e['type_name']}\n")
                f.write(f"Content: {e['content_label']}\n")
                f.write("-" * 30 + "\n")

def export_to_excel(events, output_file):
    from openpyxl.styles import Font, PatternFill
    
    # 根據日期分群
    grouped = {}
    for e in sorted(events, key=lambda x: x['id']):
        key = (e['month'], e['day'])
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(e)
            
    rows = []
    for (month, day), day_events in grouped.items():
        # 2. 區域與歌曲整合及明細提取
        # 我們需要針對每一天，辨識出不同的「區域/類別」，並將相關的「新區域」狀態與「新曲」清單整合
        
        # 暫存當天各區域的資料: key=CategoryName, value={is_new_area: bool, songs: []}
        day_categories = {}
        
        # 輔助：判斷類別與提取內容
        for e in day_events:
            name = e['name']
            if "紹介" in name or "介紹" in name: continue
            if "すたんぷカード" in name: continue # 過濾掉集章卡報酬
            if "サークルフェスタ シーズン" in name: continue # 過濾掉 Circle Festa Season
            
            # 提取歌曲 (如果有)
            extracted_songs = []
            song_match = re.search(r'《(.*?)》', name)
            if song_match:
                # 有括號，取括號內
                song_text = song_match.group(1)
                extracted_songs = [s.strip() for s in song_text.replace('、', ',').split(',')]
            elif "通常楽曲" in name:
                # 沒括號但有通常楽曲，可能是計數，或者無名單
                # 如果使用者要「只包含歌名」，這裡若無歌名則顯示概稱
                # 但如果有 《》 優先
                pass
            
            # 提取類別 (Area Name)
            # 假設格式： "類別名稱" + 空格/動詞
            # 移除常見後綴以獲得純淨名稱
            clean_name_for_cat = re.sub(r'《.*?》', '', name).strip()
            clean_name_for_cat = re.sub(r'(解禁|追加|パーフェクトチャレンジ|完全挑戰|連動楽曲|宴譜面|ウィークリーミッション).*$', '', clean_name_for_cat).strip()
            # 移除開頭的 【標籤】 (e['name']本身沒有，但在前面已被移除)
            
            # 特別處理：通常歌曲、連動、宴譜面等
            category = clean_name_for_cat
            if not category:
                if "通常楽曲" in name: category = "通常楽曲"
                elif "連動" in name: category = "連動楽曲"
                elif "宴譜面" in name: category = "宴譜面"
                elif "ミッション" in name: category = "ミッション"
                else: category = "其他"
            
            # 清理類別名稱
            category = re.sub(r'\d+曲.*$', '', category).strip() # 移除 "8曲" 等
            category = category.replace('ちほー', '').strip() # 根據使用者截圖，可能保留或移除? 截圖有 "トリコロちほー"，保留比較清楚，但使用者說「新區域」欄位
            # 使用者範例： "新區域" 欄位如果是 "トリコロちほー"，那歌曲那邊就不用重複
            
            if category not in day_categories:
                day_categories[category] = {'is_new_area': False, 'songs': []}
            
            # 標記是否為新區域解禁
            if e['content_label'] == "【新區域】" or "解禁" in name: # 簡單判斷
                day_categories[category]['is_new_area'] = True
            
            # 加入歌曲
            if extracted_songs:
                for s in extracted_songs:
                    if s not in day_categories[category]['songs']:
                        day_categories[category]['songs'].append(s)
            elif e['content_label'] == "【新曲】" and not extracted_songs:
                 # 是新曲但沒抓到歌名 (例如 "通常楽曲 1曲追加")
                 pass

        # 將整理好的當天資料轉為 Rows
        if not day_categories:
            continue
            
        for cat, data in day_categories.items():
            # 過濾掉沒內容的 (非新區域 且 無歌)
            if not data['is_new_area'] and not data['songs']:
                continue
                
            display_area = cat
            
            # --- 修改：歌曲獨立分行 ---
            # 如果有歌曲，為每一首歌產生一行
            if data['songs']:
                for song in data['songs']:
                    rows.append({
                        'month': month,
                        'day': day,
                        'area': display_area,
                        'song': song,
                        'is_new_area': data['is_new_area']
                    })
            else:
                # 只有區域沒有歌曲 (例如純解禁)
                rows.append({
                    'month': month,
                    'day': day,
                    'area': display_area,
                    'song': "",
                    'is_new_area': data['is_new_area']
                })

    # --- 雙欄佈局邏輯 (Split View) ---
    total_rows = len(rows)
    split_index = (total_rows + 1) // 2
    
    left_rows = rows[:split_index]
    right_rows = rows[split_index:]
    
    # 寫入 Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, index=False, sheet_name='Events')
        workbook = writer.book
        worksheet = writer.sheets['Events']
        
        # 定義樣式
        # 字體設定：使用微軟正黑體 (Microsoft JhengHei)
        font_name = 'Microsoft JhengHei'
        
        header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        header_font = Font(name=font_name, color='FFFFFF', bold=True, size=12)
        month_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        alignment = Alignment(wrap_text=True, vertical='center', horizontal='left', indent=0) # 改為垂直置中比較好看
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(left=Side(style='thin', color='AAAAAA'), right=Side(style='thin', color='AAAAAA'), 
                             top=Side(style='thin', color='AAAAAA'), bottom=Side(style='thin', color='AAAAAA'))

        def write_column_set(data_rows, start_col, start_row):
            # 標題
            headers = ['月份', '日期', '區域 / 類別', '新曲']
            for idx, h in enumerate(headers):
                cell = worksheet.cell(row=start_row, column=start_col + idx)
                cell.value = h
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            current_row = start_row + 1
            
            # 合併控制變數
            current_month = None
            month_start_row = current_row
            
            current_day_key = None # (Month, Day)
            day_start_row = current_row
            
            current_area_key = None # (Month, Day, Area)
            area_start_row = current_row
            
            for row_data in data_rows:
                # 寫入資料
                # 月份
                cell_month = worksheet.cell(row=current_row, column=start_col)
                cell_month.value = row_data['month']
                cell_month.border = thin_border
                cell_month.alignment = center_alignment
                
                # 日期
                cell_day = worksheet.cell(row=current_row, column=start_col + 1)
                cell_day.value = row_data['day']
                cell_day.border = thin_border
                cell_day.alignment = center_alignment
                
                # 區域
                cell_area = worksheet.cell(row=current_row, column=start_col + 2)
                cell_area.value = row_data['area']
                if row_data.get('is_new_area'):
                    cell_area.font = Font(name=font_name, bold=True, size=11)
                else:
                    cell_area.font = Font(name=font_name, size=11)
                cell_area.border = thin_border
                cell_area.alignment = alignment
                
                # 新曲
                cell_song = worksheet.cell(row=current_row, column=start_col + 3)
                cell_song.value = row_data['song']
                cell_song.border = thin_border
                cell_song.alignment = alignment
                cell_song.font = Font(name=font_name, size=11)
                
                # --- 合併邏輯檢查 ---
                # 1. 月份
                if row_data['month'] != current_month:
                    # 結算上一個月
                    if current_month is not None:
                        if current_row - 1 > month_start_row:
                            worksheet.merge_cells(start_row=month_start_row, start_column=start_col, 
                                                end_row=current_row-1, end_column=start_col)
                        for r in range(month_start_row, current_row):
                            cell = worksheet.cell(row=r, column=start_col)
                            cell.fill = month_fill
                            cell.font = Font(name=font_name, size=11)
                    current_month = row_data['month']
                    month_start_row = current_row
                
                # 2. 日期 (必須在同月的前提下，其實key已經包含month所以夠了)
                val_day_key = (row_data['month'], row_data['day'])
                if val_day_key != current_day_key:
                    if current_day_key is not None and current_row - 1 > day_start_row:
                        worksheet.merge_cells(start_row=day_start_row, start_column=start_col+1, 
                                            end_row=current_row-1, end_column=start_col+1)
                    current_day_key = val_day_key
                    day_start_row = current_row
                
                # 3. 區域 (同月份、同日期、同區域)
                val_area_key = (row_data['month'], row_data['day'], row_data['area'])
                if val_area_key != current_area_key:
                    if current_area_key is not None and current_row - 1 > area_start_row:
                        worksheet.merge_cells(start_row=area_start_row, start_column=start_col+2, 
                                            end_row=current_row-1, end_column=start_col+2)
                    current_area_key = val_area_key
                    area_start_row = current_row

                # 確保未合併單元格字體 (這段其實被上面的賦值覆蓋了，但為了保險)
                if not row_data.get('is_new_area'):
                     # cell_area 已在上面設定
                     pass
                # 日期和月份的字體也要設
                cell_day.font = Font(name=font_name, size=11)
                
                current_row += 1
            
            # --- 迴圈結束後的收尾合併 ---
            # 月份
            if current_month is not None:
                if current_row - 1 > month_start_row:
                    worksheet.merge_cells(start_row=month_start_row, start_column=start_col, 
                                        end_row=current_row-1, end_column=start_col)
                for r in range(month_start_row, current_row):
                    cell = worksheet.cell(row=r, column=start_col)
                    cell.fill = month_fill
                    cell.font = Font(name=font_name, size=11)
            
            # 日期
            if current_day_key is not None and current_row - 1 > day_start_row:
                worksheet.merge_cells(start_row=day_start_row, start_column=start_col+1, 
                                    end_row=current_row-1, end_column=start_col+1)
            
            # 區域
            if current_area_key is not None and current_row - 1 > area_start_row:
                worksheet.merge_cells(start_row=area_start_row, start_column=start_col+2, 
                                    end_row=current_row-1, end_column=start_col+2)

        write_column_set(left_rows, 1, 1)
        if right_rows:
            write_column_set(right_rows, 6, 1)
        
        # 欄寬
        for col_idx in [1, 6]: worksheet.column_dimensions[chr(64 + col_idx)].width = 6
        for col_idx in [2, 7]: worksheet.column_dimensions[chr(64 + col_idx)].width = 12
        for col_idx in [3, 8]: worksheet.column_dimensions[chr(64 + col_idx)].width = 25 # 區域稍微寬一點
        for col_idx in [4, 9]: worksheet.column_dimensions[chr(64 + col_idx)].width = 40 # 歌名
        
        worksheet.column_dimensions['E'].width = 2
        
        # 設定字體 (雙重保險，遍歷所有設定一次)
        for row in worksheet.iter_rows():
            for cell in row:
                current_font = cell.font
                new_font = Font(name=font_name, size=current_font.size, bold=current_font.b, color=current_font.color)
                cell.font = new_font

# 主程式
output_type = input("請輸入輸出類型(1: 簡要描述, 2: 完整信息, 3: Excel 導出): ")
event_folder = input('請輸入事件資料夾的路徑: ').strip('"')
default_ext = "xlsx" if output_type == '3' else "txt"
output_file = input(f'請輸入輸出文件的名稱({default_ext}): ').strip('"')

events_data = parse_event_data(event_folder)

if output_type in ['1', '2']:
    export_to_txt(events_data, output_file+".txt", output_type)
    print(f'輸出已寫入 {output_file}')
elif output_type == '3':
    export_to_excel(events_data, output_file+".xlsx")
    print(f'Excel 輸出已完成：{output_file}')
else:
    print("無效的輸出類型")
